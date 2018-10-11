import openpyxl

def write_create_doc(table_name, write_list, pk_list, fk_list, option_dict):
    """
    Creat文出力処理
    """
    file_name = table_name + ".sql" 
    with open(file_name, "w", encoding="utf-8") as f :

        f.write("create table {}(".format(table_name))
        f.write('\n')

        write_list_list = []
        for line in write_list:
            line_str = "    " + line['name'] + " " + line['field']

            if "size" in line:
                line_str = line_str + "({})".format(str(line['size']))

            if "requir" in line:
                line_str = line_str + " not null"

            if "option" in line:
                line_str = line_str + " " + line['option']

            write_list_list.append(line_str)
        
        if len(pk_list) > 0:
            write_list_list.append("    primary key({})".format(",".join(pk_list)))

        if len(fk_list) > 0:
            write_list_list.append(",\n".join(fk_list))

        if "distkey" in option_dict:
            distkey_list = option_dict["distkey"]
            write_list_list.append("    distkey({})".format(",".join(distkey_list)))

        if "sortkey" in option_dict:
            sort_list = option_dict["sortkey"]
            write_list_list.append("    compound sortkey({})".format(",".join(sort_list)))


        f.write(", \n".join(write_list_list))
        f.write("\n")

        f.write("); \n")

def read_table_sheet(sheet):
    """
    テーブル情報取得(シート単位)
    """

    # テーブル名称取得
    table_name = sheet.cell(row=2, column=6).value

    write_list = []
    option_dict = {}
    pk_list = []
    fk_list = []

    # 6行目から記載無行まで行単位で処理
    for row in sheet.iter_rows(min_row=6):
        row_dict = dict()

        if row[1].value is not None:
            row_dict['name'] = row[1].value
        else:
            continue

        if row[2].value is not None:
            row_dict['field'] = row[2].value

        if row[3].value is not None and not row[3].value == "-":
            row_dict['size'] = row[3].value

        if row[4].value is not None and row[4].value == "○":
            pk_list.append(row[1].value)

        if row[5].value is not None:
            fk_str = "    foreign key({}) references {}".format(row[1].value, row[5].value)
            fk_list.append(fk_str)

        if row[6].value is not None and row[6].value == "○":
            row_dict['requir'] = True

        if row[7].value is not None:
            for option in row[7].value.split(','):

                if option.find('encode') > -1:
                    row_dict['option'] = option
                else:
                    if option in option_dict:
                        option_dict[option].append(row[1].value)
                    else:
                        option_list = [row[1].value]
                        option_dict[option] = option_list

        write_list.append(row_dict)

    return table_name, write_list, pk_list, fk_list, option_dict


def make_create_sheet(sheet):
    """
    シートからCreate文作成
    """
    table_name, write_list, pk_list, fk_list, option_dict =  read_table_sheet(sheet)
    write_create_doc(table_name, write_list, pk_list, fk_list, option_dict)


def main():
    """
    メイン処理
    """
    # ワークシート読み込み
    wb = openpyxl.load_workbook('TableDesignDocument.xlsx')

    for name in wb.sheetnames:
        if name=="テーブル一覧":
            continue

        # 処理対象シート取得
        sheet = wb[name]
        make_create_sheet(sheet)
 
if __name__ == "__main__":
    main()
